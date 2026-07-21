"""Convert the offline MNH DHIS2 mapping workbook into deterministic JSON."""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mnid.dhis2.exceptions import DHIS2MappingError

UID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9]{10}$")
OPERAND_RE = re.compile(r"^[A-Za-z][A-Za-z0-9]{10}(?:\.[A-Za-z][A-Za-z0-9]{10})?$")
DEFAULT_SHEET_HINTS = ("mnh indicators", "indicators", "mapping")


def stable_id(value: str) -> str:
    """Create a stable lowercase identifier from an indicator name."""
    normalized = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    if not normalized:
        raise DHIS2MappingError("Indicator name cannot produce a stable ID")
    return normalized


def normalize_heading(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def operand(dx: str) -> dict[str, Any]:
    value = dx.strip()
    if not OPERAND_RE.fullmatch(value):
        raise DHIS2MappingError(f"Invalid DHIS2 operand {value!r}")
    parts = value.split(".", 1)
    return {
        "type": "data_element_operand" if len(parts) == 2 else "data_element",
        "dx": value,
        "data_element_id": parts[0],
        "category_option_combo_id": parts[1] if len(parts) == 2 else None,
    }


def _atomic_values(text: str) -> list[str]:
    return re.findall(r"[A-Za-z][A-Za-z0-9]{10}(?:\.[A-Za-z][A-Za-z0-9]{10})?", text or "")


def _resolve_reference(name: str, names_to_ids: dict[str, str]) -> str | None:
    return names_to_ids.get(re.sub(r"\s+", " ", name.strip()).casefold())


def _calculation(
    mapping: str | None,
    formula: str | None,
    names_to_ids: dict[str, str],
) -> tuple[dict[str, Any] | None, str, list[str]]:
    raw_mapping = str(mapping or "").strip()
    raw_formula = str(formula or "").strip()
    ids = _atomic_values(raw_mapping)
    messages: list[str] = []
    if not raw_mapping and raw_formula:
        terms = [term.strip() for term in raw_formula.lstrip("=").split("+") if term.strip()]
        refs = [_resolve_reference(term, names_to_ids) for term in terms]
        if len(terms) >= 2 and all(refs):
            return {"operation": "sum_indicators", "indicator_ids": refs}, "valid", messages
        return None, "review_required", ["Formula references could not be resolved unambiguously."]
    if not ids:
        return None, "review_required", ["No valid DHIS2 mapping was found in the source workbook."]

    if "/" in raw_mapping and len(ids) == 2:
        return {
            "operation": "percentage",
            "multiplier": 100,
            "numerator": {"operands": [operand(ids[0])]},
            "denominator": {"operands": [operand(ids[1])]},
            "zero_denominator_result": None,
        }, "valid", messages

    if "-" in raw_mapping and len(ids) >= 2:
        if "livebirth" in raw_formula.lower() or "live birth" in raw_formula.lower():
            calc = {
                "operation": "subtract",
                "minuend": {"indicator_ids": ["live_births"]},
                "subtrahend": {"operands": [operand(item) for item in ids[-1:]]},
            }
        else:
            calc = {
                "operation": "subtract",
                "minuend": {"operands": [operand(ids[0])]},
                "subtrahend": {"operands": [operand(item) for item in ids[1:]]},
            }
        return calc, "review_required", [
            "Workbook subtraction was preserved as a count; confirm the intended dashboard value type."
        ]

    ops = [operand(item) for item in ids]
    return {
        "operation": "direct" if len(ops) == 1 else "sum",
        "operands": ops,
    }, "valid", messages


def _choose_sheet(workbook, requested: str | None):
    if requested:
        if requested not in workbook.sheetnames:
            raise DHIS2MappingError(f"Worksheet {requested!r} was not found")
        return workbook[requested]
    for name in workbook.sheetnames:
        if normalize_heading(name).replace("_", " ") in DEFAULT_SHEET_HINTS:
            return workbook[name]
    return workbook[workbook.sheetnames[0]]


def _read_existing(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
        return loaded if isinstance(loaded, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def convert_workbook(
    input_path: Path,
    output_path: Path,
    *,
    sheet: str | None = None,
    mapping_version: str = "2026-07-20",
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Convert all non-empty workbook rows and return mapping plus change report."""
    if not input_path.exists():
        raise DHIS2MappingError(f"Workbook not found: {input_path}")
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=False)
    except Exception as exc:
        raise DHIS2MappingError(f"Unable to read workbook: {input_path.name}") from exc
    ws = _choose_sheet(workbook, sheet)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise DHIS2MappingError("Workbook worksheet is empty")
    headings = [normalize_heading(value) for value in rows[0]]
    required = "indicator_name"
    if required not in headings:
        raise DHIS2MappingError("Workbook must contain an INDICATOR NAME column")
    index = {name: pos for pos, name in enumerate(headings) if name}
    existing = _read_existing(output_path)
    old_indicators = {item.get("id"): item for item in existing.get("indicators", []) if item.get("id")}

    staged: list[tuple[int, str, str | None, str | None, str | None]] = []
    for row_number, row in enumerate(rows[1:], 2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        def get(*names: str) -> str | None:
            for name in names:
                pos = index.get(name)
                if pos is not None and pos < len(row) and row[pos] is not None:
                    value = str(row[pos]).strip()
                    if value:
                        return value
            return None
        name = get("indicator_name", "name")
        if not name:
            name = f"Unnamed workbook row {row_number}"
        explicit_id = get("internal_id", "indicator_id", "id")
        mapping = get("dhis2_id", "dhis2_ids", "mapping")
        formula = get("indicator_calculation", "calculation", "formula")
        staged.append((row_number, name, explicit_id, mapping, formula))

    names_to_ids = {
        re.sub(r"\s+", " ", name.strip()).casefold(): explicit_id or stable_id(name)
        for _, name, explicit_id, _, _ in staged
    }
    indicators: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, name, explicit_id, raw_mapping, formula in staged:
        generated = explicit_id or stable_id(name)
        existing_match = old_indicators.get(generated)
        indicator_id = existing_match.get("id") if existing_match else generated
        if indicator_id in seen:
            raise DHIS2MappingError(f"Duplicate indicator ID {indicator_id!r}")
        seen.add(indicator_id)
        calculation, status, messages = _calculation(raw_mapping, formula, names_to_ids)
        value_type = "percentage" if calculation and calculation["operation"] == "percentage" else (
            "unknown" if calculation is None else "count"
        )
        enabled = calculation is not None and status == "valid"
        if calculation and calculation["operation"] == "subtract":
            enabled = True
        indicators.append({
            "id": indicator_id,
            "name": name,
            "description": existing_match.get("description") if existing_match else None,
            "source": "dhis2",
            "enabled": enabled,
            "value_type": value_type,
            "calculation": calculation,
            "validation": {"status": status, "messages": messages},
            "source_reference": {
                "workbook_row": row_number,
                "original_indicator_name": name,
                "original_mapping": raw_mapping or formula,
            },
        })

    removed = sorted(set(old_indicators) - seen)
    for indicator_id in removed:
        retained = dict(old_indicators[indicator_id])
        retained["enabled"] = False
        validation = dict(retained.get("validation") or {})
        validation["status"] = "review_required"
        validation["messages"] = list(validation.get("messages") or []) + ["Removed from the current workbook; retained disabled."]
        retained["validation"] = validation
        indicators.append(retained)

    result = {
        "schema_version": "1.0",
        "mapping_version": mapping_version,
        "source": {
            "name": "Malawi HMIS DHIS2",
            "base_url": "https://dhis2.health.gov.mw",
            "source_workbook": input_path.name,
            "worksheet": ws.title,
        },
        "reporting_period": {
            "period_type": "monthly", "start_period": "202504", "end_period": "202605"
        },
        "indicators": indicators,
    }
    # Validate the complete candidate before any atomic replacement. Import here
    # keeps the standalone conversion primitives reusable without import cycles.
    from mnid.dhis2.schemas import validate_indicator_mapping
    validate_indicator_mapping(result)
    old_by_id = old_indicators
    new_by_id = {item["id"]: item for item in indicators}
    added = [key for key in new_by_id if key not in old_by_id]
    unchanged = [key for key in new_by_id if key in old_by_id and new_by_id[key] == old_by_id[key]]
    updated = [key for key in new_by_id if key in old_by_id and new_by_id[key] != old_by_id[key]]
    all_dx = sorted({op["dx"] for item in indicators for op in _walk_operands(item.get("calculation"))})
    statuses = [item["validation"]["status"] for item in indicators]
    operations = [item["calculation"]["operation"] for item in indicators if item.get("calculation")]
    report = {
        "status": "success",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_workbook": input_path.name,
        "worksheet": ws.title,
        "mapping_version": mapping_version,
        "rows_read": len(staged),
        "indicators_generated": len(indicators),
        "enabled_indicators": sum(bool(item["enabled"]) for item in indicators),
        "disabled_indicators": sum(not item["enabled"] for item in indicators),
        "valid_indicators": statuses.count("valid"),
        "review_required": statuses.count("review_required"),
        "rejected_indicators": statuses.count("rejected"),
        "unique_dx_operands": len(all_dx),
        "direct_indicators": operations.count("direct"),
        "sum_indicators": operations.count("sum"),
        "subtraction_indicators": operations.count("subtract"),
        "percentage_indicators": operations.count("percentage"),
        "derived_indicators": operations.count("sum_indicators"),
        "changes": {
            "added": added, "updated": updated, "unchanged": unchanged,
            "disabled": [item["id"] for item in indicators if not item["enabled"]],
            "removed_from_workbook": removed,
            "mapping_changed": updated, "formula_changed": [],
        },
        "warnings": [message for item in indicators for message in item["validation"]["messages"]],
        "errors": [],
    }
    return result, report


def _walk_operands(value: Any):
    if isinstance(value, dict):
        if "dx" in value:
            yield value
        for child in value.values():
            yield from _walk_operands(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk_operands(child)


def _encoded(value: dict[str, Any]) -> bytes:
    return (json.dumps(value, ensure_ascii=False, indent=4) + "\n").encode("utf-8")


def atomic_write_if_changed(path: Path, value: dict[str, Any]) -> bool:
    """Atomically write deterministic JSON, preserving unchanged files."""
    payload = _encoded(value)
    if path.exists() and path.read_bytes() == payload:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except OSError:
            pass
        raise
    return True


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--mapping-version", default="2026-07-20")
    parser.add_argument("--validate-only", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--strict", action="store_true")
    parser.add_argument("--report-output", type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    report_path = args.report_output or args.output.with_name("indicator_conversion_report.json")
    try:
        result, report = convert_workbook(
            args.input, args.output, sheet=args.sheet, mapping_version=args.mapping_version
        )
        if args.strict and (report["review_required"] or report["rejected_indicators"]):
            raise DHIS2MappingError("Strict conversion rejected mappings requiring review")
        if not args.validate_only:
            atomic_write_if_changed(args.output, result)
            atomic_write_if_changed(report_path, report)
        print(
            f"rows={report['rows_read']} indicators={report['indicators_generated']} "
            f"enabled={report['enabled_indicators']} review={report['review_required']} "
            f"rejected={report['rejected_indicators']} unique_dx={report['unique_dx_operands']}"
        )
        return 0
    except DHIS2MappingError as exc:
        print(f"Conversion failed: {exc}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
