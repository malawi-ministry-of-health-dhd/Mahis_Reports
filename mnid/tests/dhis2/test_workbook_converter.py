"""Workbook conversion tests using generated temporary workbooks."""

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from mnid.dhis2.exceptions import DHIS2MappingError
from mnid.dhis2.tools.convert_indicator_workbook import atomic_write_if_changed, convert_workbook


class ConverterTests(unittest.TestCase):
    def make_workbook(self, directory: Path) -> Path:
        path = directory / "mapping.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "MNH indicators"
        ws.append(["INDICATOR NAME", "DHIS2 ID", "Indicator Calculation"])
        ws.append(["Live Births", "iBBnHx1Uf50, EeywK6AHQdK", None])
        ws.append(["Stillbirths", "YLydKJIyGFF", None])
        ws.append(["Total Births", None, "Live Births + Stillbirths"])
        ws.append(["Coverage", "n7XnfWv0tHo/RS3PtRniRns.JPgETYeD4qF", "numerator/denominator"])
        ws.append(["Missing", None, None])
        wb.save(path)
        return path

    def test_direct_sum_derived_percentage_and_missing(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            result, report = convert_workbook(self.make_workbook(root), root / "out.json")
            by_id = {item["id"]: item for item in result["indicators"]}
            self.assertEqual("sum", by_id["live_births"]["calculation"]["operation"])
            self.assertEqual("direct", by_id["stillbirths"]["calculation"]["operation"])
            self.assertEqual("sum_indicators", by_id["total_births"]["calculation"]["operation"])
            self.assertEqual("percentage", by_id["coverage"]["calculation"]["operation"])
            self.assertFalse(by_id["missing"]["enabled"])
            self.assertEqual(5, report["rows_read"])

    def test_deterministic_atomic_write(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "out.json"
            value = {"b": 1, "a": [2]}
            self.assertTrue(atomic_write_if_changed(path, value))
            first = path.read_bytes()
            self.assertFalse(atomic_write_if_changed(path, value))
            self.assertEqual(first, path.read_bytes())
            self.assertEqual(value, json.loads(first))

    def test_missing_workbook_and_duplicate_id_fail(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            with self.assertRaises(DHIS2MappingError):
                convert_workbook(root / "missing.xlsx", root / "out.json")
            path = self.make_workbook(root)
            wb = load_workbook_compat(path)
            wb.active.append(["Live-Births", "WjHvEHMCyKo", None])
            wb.save(path)
            with self.assertRaises(DHIS2MappingError):
                convert_workbook(path, root / "out.json")


def load_workbook_compat(path: Path):
    from openpyxl import load_workbook
    return load_workbook(path)


if __name__ == "__main__":
    unittest.main()
